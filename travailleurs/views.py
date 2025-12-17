from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from .models import Travailleurs
from .forms import TravailleursForm
import csv
import openpyxl
from django.shortcuts import render
from django.contrib.auth.decorators import login_required

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import Travailleurs
from django.db.models import Count
from django.shortcuts import render
from .models import Travailleurs
from collections import defaultdict


@login_required
def dashboard(request):
    sections = ['Whiskers','Laser', 'Handsanding', 'Grinding', 'Spray','Sample','Out Going','Mechanics','Responsible']
    statuts = ['CDI', 'CDD', 'Temporaire', 'Journalier']

    table_data = []
    for statut in statuts:
        row = {'statut': statut, 'total': 0}
        for section in sections:
            count = Travailleurs.objects.filter(statut=statut, section=section,etat='Actif').count()
            row[section] = count
            row['total'] += count
        table_data.append(row)

    total_par_section = {section: sum(row[section] for row in table_data) for section in sections}
    total_general = sum(row['total'] for row in table_data)

    total_travailleurs = Travailleurs.objects.count()
    actifs = Travailleurs.objects.filter(etat='Actif').count()
    inactifs = Travailleurs.objects.filter(etat='Inactif').count()

    return render(request, 'travailleurs/dashboard.html', {
        'sections': sections,
        'table_data': table_data,
        'total_par_section': total_par_section,
        'total_general': total_general,
        'total_travailleurs': total_travailleurs,
        'actifs': actifs,
        'inactifs': inactifs,
        'user': request.user
    })



from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages

def user_login(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, "Nom d'utilisateur ou mot de passe incorrect")
    return render(request, 'travailleurs/login.html')



# Liste des travailleurs avec recherche
from .models import Travailleurs  # assure-toi que c'est singulier

from django.shortcuts import render

def travailleur_list(request):
    query = request.GET.get('q')  # ce que l'utilisateur tape dans la barre de recherche
    if query:
        travailleurs = Travailleurs.objects.filter(
            matricule__icontains=query,  # recherche par matricule
            etat="Actif"
        )
    else:
        travailleurs = Travailleurs.objects.filter(etat="Actif")
    
    return render(request, 'travailleurs/travailleur_list.html', {'travailleurs': travailleurs})

# Log out
def logout(request):
    return render(request, 'travailleurs/logout.html')

# Détails d'un travailleur
def travailleur_detail(request, pk):
    travailleur = get_object_or_404(Travailleurs, pk=pk)
    return render(request, 'travailleurs/travailleur_detail.html', {'travailleur': travailleur})

# Créer un travailleur
def travailleur_create(request):
    if request.method == "POST":
        form = TravailleursForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('travailleur_list')
    else:
        form = TravailleursForm()
    return render(request, 'travailleurs/travailleur_form.html', {'form': form})

# Modifier un travailleur
def travailleur_update(request, pk):
    travailleur = get_object_or_404(Travailleurs, pk=pk)
    if request.method == "POST":
        form = TravailleursForm(request.POST, instance=travailleur)
        if form.is_valid():
            form.save()
            return redirect('travailleur_list')
    else:
        form = TravailleursForm(instance=travailleur)
    return render(request, 'travailleurs/travailleur_form.html', {'form': form})

# Supprimer un travailleur
def travailleur_delete(request, pk):
    travailleur = get_object_or_404(Travailleurs, pk=pk)
    if request.method == "POST":
        travailleur.delete()
        return redirect('travailleur_list')
    return render(request, 'travailleurs/travailleur_confirm_delete.html', {'travailleur': travailleur})


# Exporter les travailleurs actifs en Excel
def export_travailleurs_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Travailleurs'

    headers = ['Matricule','Nom','Prénom','Adresse','Section','Fonction','Statut','État','Date Embauche','Date Débauche','Motif Débauche']
    sheet.append(headers)

    travailleurs = Travailleurs.objects.filter(etat="Actif")
    for t in travailleurs:
        sheet.append([
            t.matricule, t.nom, t.prenom, t.adresse, t.section, t.fonction,
            t.statut, t.etat, t.date_embauche, t.date_debauche, t.motif_de_debauche
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="travailleurs.xlsx"'
    workbook.save(response)
    return response

from .models import Absence
from .forms import AbsenceForm
def liste_absences(request):
    absences = Absence.objects.all().order_by('-date')

    date_debut = request.GET.get("date_debut")
    date_fin = request.GET.get("date_fin")

    if date_debut and date_fin:
        absences = absences.filter(date__range=[date_debut, date_fin])

    return render(request, "travailleurs/liste_absences.html", {
        "absences": absences,
    })

def modifier_absence(request, pk):
    absence = get_object_or_404(Absence, pk=pk)
    if request.method == "POST":
        form = AbsenceForm(request.POST, instance=absence)
        if form.is_valid():
            form.save()
            return redirect('liste_absences')
    else:
        form = AbsenceForm(instance=absence)
    return render(request, 'travailleurs/form_absence.html', {'form': form})


def supprimer_absence(request, pk):
    absence = get_object_or_404(Absence, pk=pk)
    if request.method == "POST":
        absence.delete()
        return redirect('liste_absences')
    return render(request, 'travailleurs/confirmer_supprimer.html', {'absence': absence})

def creer_absence(request):
    if request.method == "POST":
        form = AbsenceForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('liste_absences')
    else:
        form = AbsenceForm()
    return render(request, 'travailleurs/form_absence.html', {'form': form})




# eto ny code momba ny Nightshift rehetra


from .forms import NightShiftForm
from .models import NightShift
def creer_nightshift(request):
    if request.method == 'POST':
        form = NightShiftForm(request.POST)
        if form.is_valid():
            date_debut = form.cleaned_data['date_debut']
            date_fin = form.cleaned_data['date_fin']
            travailleurs = form.cleaned_data['travailleurs']

            # Supprimer anciens enregistrements de la période
           #  NightShift.objects.filter(date_debut=date_debut, date_fin=date_fin).delete()

            # Créer les nouveaux
            for t in travailleurs:
                NightShift.objects.create(
                    travailleur=t,
                    date_debut=date_debut,
                    date_fin=date_fin
                )
            return redirect('liste_nightshift')
    else:
        form = NightShiftForm()

    return render(request, 'travailleurs/creer_nightshift.html', {'form': form})

from .models import NightShift
from .forms import NightShiftForm



def liste_nightshift(request):
    nightshifts = NightShift.objects.all()
    return render(request, 'travailleurs/liste_nightshift.html', {'nightshifts': nightshifts})


def supprimer_nightshift(request, pk):
    nightshift = get_object_or_404(NightShift, pk=pk)
    if request.method == 'POST':
        nightshift.delete()
        return redirect('liste_nightshift')
    return render(request, 'travailleurs/supprimer_nightshift.html', {'nightshift': nightshift})


from datetime import datetime, timedelta
from .models import Travailleurs, Absence, NightShift
from .forms import PeriodeForm
import calendar

def etat_presence(request):
    table = []
    dates = []
    
    if request.method == "POST":
        form = PeriodeForm(request.POST)
        if form.is_valid():
            date_debut = form.cleaned_data['date_debut']
            date_fin = form.cleaned_data['date_fin']

            # Générer toutes les dates entre debut et fin
            delta = date_fin - date_debut
            dates = [date_debut + timedelta(days=i) for i in range(delta.days + 1)]

            travailleurs = Travailleurs.objects.filter(etat='Actif')

            for t in travailleurs:
                row = {'travailleur': t, 'jours': []}
                for d in dates:
                    # NightShift
                    if NightShift.objects.filter(travailleur=t, date_debut__lte=d, date_fin__gte=d).exists():
                        row['jours'].append('NS')
                        continue

                    # Absences
                    absences_obj = Absence.objects.filter(travailleur=t, date=d)
                    if absences_obj.exists():
                        a = absences_obj.first()
                        if a.type_absence == 'jour':
                            row['jours'].append(a.motif.capitalize())
                        else:  # heure
                            if d.weekday() < 5:
                                row['jours'].append(str(10 - a.duree))
                            elif d.weekday() == 5:
                                row['jours'].append(str(8 - a.duree))
                            else:
                                row['jours'].append(0)
                        continue

                    # Jour normal
                    if d.weekday() < 5:
                        row['jours'].append(10)
                    elif d.weekday() == 5:
                        row['jours'].append(8)
                    else:
                        row['jours'].append(0)

                table.append(row)
    else:
        form = PeriodeForm()

    absences = ['Permission','Repos','Congé','Sansmotif','Autre']

    return render(request, 'travailleurs/etat_presence.html', {
        'table': table,
        'dates': dates,
        'absences': absences,
        'form': form,  # 🔹 ajouter le formulaire ici
    })

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from datetime import timedelta
from .models import Travailleurs, NightShift, Absence


def export_presence_excel(request):
    # Récupération des paramètres GET
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')

    if not date_debut or not date_fin:
        return HttpResponse("Veuillez générer un état d'abord.", status=400)

    date_debut = datetime.strptime(date_debut, "%Y-%m-%d").date()
    date_fin = datetime.strptime(date_fin, "%Y-%m-%d").date()

    # Générer les dates
    delta = date_fin - date_debut
    dates = [date_debut + timedelta(days=i) for i in range(delta.days + 1)]

    travailleurs = Travailleurs.objects.filter(etat='Actif')

    # Créer un fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "État de présence"

    # En-têtes
    ws.append(["Matricule", "Nom"] + [d.strftime("%d/%m") for d in dates])

    # Style en-têtes
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="333333")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Remplir le tableau
    for t in travailleurs:
        row = [t.matricule, f"{t.nom} {t.prenom}"]
        for d in dates:
            # Vérifier NightShift
            if NightShift.objects.filter(travailleur=t, date_debut__lte=d, date_fin__gte=d).exists():
                row.append("NS")
                continue

            # Vérifier Absence
            absences = Absence.objects.filter(travailleur=t, date=d)
            if absences.exists():
                a = absences.first()
                if a.type_absence == "jour":
                    row.append(a.motif.capitalize())
                else:
                    if d.weekday() < 5:  # Lundi-Vendredi
                        row.append(10 - a.duree)
                    elif d.weekday() == 5:  # Samedi
                        row.append(8 - a.duree)
                    else:
                        row.append(0)
                continue

            # Jour normal
            if d.weekday() < 5:
                row.append(10)
            elif d.weekday() == 5:
                row.append(8)
            else:
                row.append(0)

        ws.append(row)

    # Auto-ajustement largeur colonnes
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="etat_presence_{date_debut}_{date_fin}.xlsx"'

    wb.save(response)
    return response


def export_absences_excel(request):
    absences = Absence.objects.all()
    date_debut = request.GET.get("date_debut")
    date_fin = request.GET.get("date_fin")

    if date_debut and date_fin:
        absences = absences.filter(date__range=[date_debut, date_fin])

    # Création du fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Absences"

    headers = ["Matricule", "Nom et Prenom", "Date", "Motif", "Type", "Durée"]
    ws.append(headers)

    for a in absences:
        ws.append([
            a.travailleur.matricule,
            f"{a.travailleur.nom} {a.travailleur.prenom}",
            a.date.strftime("%d/%m/%Y"),
            a.motif,
            a.type_absence,
            a.duree,
        ])

    response = HttpResponse(
        content=openpyxl.writer.excel.save_virtual_workbook(wb),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=absences.xlsx'
    return response


import openpyxl
from io import BytesIO
from django.http import HttpResponse
from .models import Absence

def export_absences_excel(request):
    absences = Absence.objects.all()
    date_debut = request.GET.get("date_debut")
    date_fin = request.GET.get("date_fin")

    if date_debut and date_fin:
        absences = absences.filter(date__range=[date_debut, date_fin])

    # Création du fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Absences"

    headers = ["Matricule", "Nom et Prenom", "Date", "Motif", "Type", "Durée"]
    ws.append(headers)

    for a in absences:
        ws.append([
            a.travailleur.matricule,
            f"{a.travailleur.nom} {a.travailleur.prenom}",
            a.date.strftime("%d/%m/%Y"),
            a.motif,
            a.type_absence,
            a.duree,
        ])

    # Écriture en mémoire
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=absences.xlsx'
    return response
